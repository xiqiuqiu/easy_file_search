import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
import csv
import subprocess
import platform

def search_excel(file_path, keywords):
    matches = []
    keyword_counts = {keyword: 0 for keyword in keywords}  # 初始化所有关键词计数
    
    try:
        wb = load_workbook(file_path, data_only=True)
        
        # 第一遍：收集匹配项并统计关键词总数
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_str = str(cell.value)
                        for keyword in keywords:
                            if keyword in cell_str:
                                # 统计关键词在当前单元格中的出现次数
                                count_in_cell = cell_str.count(keyword)
                                keyword_counts[keyword] += count_in_cell
                                
                                # 暂时只存储基本匹配信息
                                matches.append([
                                    os.path.basename(file_path),
                                    sheet_name,
                                    cell.coordinate,
                                    keyword,  # 暂时只存储关键词
                                    cell_str,
                                    file_path
                                ])
        wb.close()
    except Exception as e:
        print(f"[错误] 处理文件失败：{file_path}，原因：{e}")
    
    # 第二遍：为每个匹配项添加完整的统计信息
    final_matches = []
    for match in matches:
        keyword = match[3]
        total_count = keyword_counts[keyword]
        updated_match = match.copy()
        updated_match[3] = f"{keyword} (本文件共{total_count}次)"
        final_matches.append(updated_match)
    
    # 返回匹配结果和该文件的关键词统计
    return final_matches, keyword_counts

def search_all_excels(directory, keywords):
    results = []
    global_keyword_counts = {keyword: 0 for keyword in keywords}  # 全局关键词统计
    
    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.xlsx'):
                file_path = os.path.join(root, file)
                file_results, file_keyword_counts = search_excel(file_path, keywords)
                results += file_results
                
                # 累加到全局统计
                for keyword, count in file_keyword_counts.items():
                    global_keyword_counts[keyword] += count
    
    return results, global_keyword_counts

def browse_directory():
    path = filedialog.askdirectory()
    if path:
        path_var.set(path)

# 添加全局變量來存儲當前搜索結果
current_search_results = []

def start_search():
    global current_search_results
    directory = path_var.get()
    keyword_input = keywords_var.get()
    if not os.path.exists(directory):
        messagebox.showerror("错误", "请选择有效的文件夹路径")
        return
    if not keyword_input.strip():
        messagebox.showerror("错误", "请输入至少一个关键词")
        return
    keywords = [kw.strip() for kw in keyword_input.split(",") if kw.strip()]
    results, total_stats = search_all_excels(directory, keywords)  # 现在直接返回统计信息
    current_search_results = results  # 保存完整的搜索結果

    # 清空之前的结果
    for row in tree.get_children():
        tree.delete(row)
    
    # 检查是否所有结果都来自同一个文件
    unique_files = set(row[5] for row in results)  # 文件路径在索引5
    is_single_file = len(unique_files) == 1
    
    # 重新配置树视图列
    if is_single_file and results:
        # 单文件模式：隐藏文件名和文件路径列，在状态栏显示文件信息
        file_path = list(unique_files)[0]
        file_name = os.path.basename(file_path)
        
        # 更新列配置（使用列索引而不是列名）
        tree["displaycolumns"] = (1, 2, 3, 4)  # 工作表、单元格、关键词、内容
        
        # 插入完整的数据（保持所有6列数据，但只显示指定的列）
        for row in results:
            tree.insert('', 'end', values=row)
        
        # 构建状态信息，包含统计
        stats_text = " | ".join([f"{kw}: {count}次" for kw, count in total_stats.items()])
        status_var.set(f"搜索完成，共命中 {len(results)} 条记录 | 文件：{file_name} | {stats_text}")
    else:
        # 多文件模式：显示所有列
        tree["displaycolumns"] = (0, 1, 2, 3, 4, 5)  # 所有列
        
        for row in results:
            tree.insert('', 'end', values=row)
        
        # 构建状态信息，包含统计
        stats_text = " | ".join([f"{kw}: {count}次" for kw, count in total_stats.items()])
        status_var.set(f"搜索完成，共命中 {len(results)} 条记录 | {stats_text}")

def export_csv():
    if not tree.get_children():
        messagebox.showinfo("提示", "没有可以导出的结果")
        return
    file_path = filedialog.asksaveasfilename(defaultextension=".csv",
                                             filetypes=[("CSV files", "*.csv")])
    if file_path:
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            
            # 根據當前顯示模式決定CSV標題
            displayed_columns = tree["displaycolumns"]
            if len(displayed_columns) == 4:  # 單文件模式
                writer.writerow(["工作表", "单元格", "命中关键词", "内容"])
                # 添加文件信息註釋
                status_text = status_var.get()
                if "文件：" in status_text:
                    file_name = status_text.split("文件：")[1]
                    writer.writerow([f"# 文件: {file_name}", "", "", ""])
                
                # 只導出顯示的列（工作表、單元格、关键词、內容）
                for row_id in tree.get_children():
                    values = tree.item(row_id)["values"]
                    writer.writerow([values[1], values[2], values[3], values[4]])
            else:  # 多文件模式
                writer.writerow(["文件名", "工作表", "单元格", "命中关键词", "内容", "文件路径"])
                for row_id in tree.get_children():
                    writer.writerow(tree.item(row_id)["values"])
        messagebox.showinfo("成功", f"结果已保存到 {file_path}")

def open_file():
    """打開選中行對應的文件"""
    selection = tree.selection()
    if not selection:
        messagebox.showinfo("提示", "請先選擇一條記錄")
        return
    
    # 獲取選中行的索引
    item = tree.item(selection[0])
    row_index = tree.index(selection[0])
    
    # 根據當前顯示模式確定文件路徑
    displayed_columns = tree["displaycolumns"]
    if len(displayed_columns) == 4:  # 單文件模式
        # 從全局搜索結果中獲取文件路徑（現在數據是完整的，直接從values獲取）
        values = item['values']
        file_path = values[5]  # 文件路徑在第6列（索引5）
    else:  # 多文件模式
        values = item['values']
        file_path = values[5]  # 文件路徑在第6列（索引5）
    
    if not os.path.exists(file_path):
        messagebox.showerror("錯誤", f"文件不存在：{file_path}")
        return
    
    try:
        # 根據操作系統使用不同的命令打開文件
        if platform.system() == 'Windows':
            os.startfile(file_path)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', file_path])
        else:  # Linux
            subprocess.run(['xdg-open', file_path])
    except Exception as e:
        messagebox.showerror("錯誤", f"無法打開文件：{e}")

# GUI 界面搭建
root = tk.Tk()
root.title("Excel关键词搜索工具")
root.geometry("1000x550")

path_var = tk.StringVar()
keywords_var = tk.StringVar()
status_var = tk.StringVar(value="就緒")

frame_top = tk.Frame(root)
frame_top.pack(fill="x", padx=10, pady=5)

tk.Label(frame_top, text="搜索路徑：").pack(side="left")
tk.Entry(frame_top, textvariable=path_var, width=60).pack(side="left", padx=5)
tk.Button(frame_top, text="瀏覽...", command=browse_directory).pack(side="left")

frame_mid = tk.Frame(root)
frame_mid.pack(fill="x", padx=10, pady=5)

tk.Label(frame_mid, text="关键词（多個用英文逗號分隔）：").pack(side="left")
tk.Entry(frame_mid, textvariable=keywords_var, width=50).pack(side="left", padx=5)

frame_btn = tk.Frame(root)
frame_btn.pack(fill="x", padx=10, pady=5)

tk.Button(frame_btn, text="搜索Excel", command=start_search).pack(side="left")
tk.Button(frame_btn, text="导出為CSV", command=export_csv).pack(side="left", padx=10)
tk.Button(frame_btn, text="打開文件", command=open_file).pack(side="left", padx=10)

columns = ["文件名", "工作表", "單元格", "命中关键词", "內容", "文件路徑"]
tree = ttk.Treeview(root, columns=columns, show="headings")
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=150 if col not in ["內容", "文件路徑"] else 300)

tree.pack(fill="both", expand=True, padx=10, pady=5)

status_bar = tk.Label(root, textvariable=status_var, anchor="w")
status_bar.pack(fill="x")

root.mainloop()
