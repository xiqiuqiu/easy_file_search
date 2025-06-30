import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
import csv
import subprocess
import platform

# 导入国际化支持
from i18n import init_i18n, t, set_language, get_available_languages, get_language_name, get_current_language
from config import get_config

class ExcelSearchApp:
    def __init__(self):
        # 初始化配置
        self.config = get_config()
        
        # 初始化国际化
        self.i18n = init_i18n()
        
        # 初始化主窗口
        self.root = tk.Tk()
        self.setup_variables()
        self.current_search_results = []
        self.create_ui()
        
        # 恢复上次搜索路径
        last_path = self.config.get_last_search_path()
        if last_path and os.path.exists(last_path):
            self.path_var.set(last_path)
        
    def setup_variables(self):
        """设置界面变量"""
        self.path_var = tk.StringVar()
        self.keywords_var = tk.StringVar()
        self.status_var = tk.StringVar(value=t('ready'))
        
    def create_ui(self):
        """创建用户界面"""
        self.update_window_title()
        
        # 设置窗口大小
        geometry = self.config.get_window_geometry()
        self.root.geometry(geometry)
        
        # 绑定窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # 创建菜单栏
        self.create_menu()
        
        # 创建主界面
        self.create_main_interface()
        
    def create_menu(self):
        """创建菜单栏"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # 语言菜单
        language_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=t('app_title') if get_current_language() == 'en' else "语言", menu=language_menu)
        
        for lang_code in get_available_languages():
            lang_name = get_language_name(lang_code)
            language_menu.add_command(
                label=lang_name,
                command=lambda lc=lang_code: self.change_language(lc)
            )
    
    def create_main_interface(self):
        """创建主界面"""
        # 顶部路径选择框架
        frame_top = tk.Frame(self.root)
        frame_top.pack(fill="x", padx=10, pady=5)
        
        self.path_label = tk.Label(frame_top, text=t('search_path'))
        self.path_label.pack(side="left")
        
        tk.Entry(frame_top, textvariable=self.path_var, width=60).pack(side="left", padx=5)
        
        self.browse_btn = tk.Button(frame_top, text=t('browse'), command=self.browse_directory)
        self.browse_btn.pack(side="left")
        
        # 中部关键词输入框架
        frame_mid = tk.Frame(self.root)
        frame_mid.pack(fill="x", padx=10, pady=5)
        
        self.keywords_label = tk.Label(frame_mid, text=t('keywords'))
        self.keywords_label.pack(side="left")
        
        tk.Entry(frame_mid, textvariable=self.keywords_var, width=50).pack(side="left", padx=5)
        
        # 按钮框架
        frame_btn = tk.Frame(self.root)
        frame_btn.pack(fill="x", padx=10, pady=5)
        
        self.search_btn = tk.Button(frame_btn, text=t('search_excel'), command=self.start_search)
        self.search_btn.pack(side="left")
        
        self.export_btn = tk.Button(frame_btn, text=t('export_csv'), command=self.export_csv)
        self.export_btn.pack(side="left", padx=10)
        
        self.open_btn = tk.Button(frame_btn, text=t('open_file'), command=self.open_file)
        self.open_btn.pack(side="left", padx=10)
        
        # 结果显示表格
        self.create_result_tree()
        
        # 状态栏
        self.status_bar = tk.Label(self.root, textvariable=self.status_var, anchor="w")
        self.status_bar.pack(fill="x")
    
    def create_result_tree(self):
        """创建结果显示表格"""
        columns = [t('file_name'), t('worksheet'), t('cell'), t('keyword'), t('content'), t('file_path')]
        self.tree = ttk.Treeview(self.root, columns=columns, show="headings")
        
        for col in columns:
            self.tree.heading(col, text=col)
            width = 150 if col not in [t('content'), t('file_path')] else 300
            self.tree.column(col, width=width)
        
        self.tree.pack(fill="both", expand=True, padx=10, pady=5)
    
    def change_language(self, language):
        """切换语言"""
        if set_language(language):
            self.refresh_ui()
    
    def refresh_ui(self):
        """刷新界面文本"""
        self.update_window_title()
        self.status_var.set(t('ready'))
        
        # 更新标签文本
        self.path_label.config(text=t('search_path'))
        self.keywords_label.config(text=t('keywords'))
        self.browse_btn.config(text=t('browse'))
        self.search_btn.config(text=t('search_excel'))
        self.export_btn.config(text=t('export_csv'))
        self.open_btn.config(text=t('open_file'))
        
        # 更新表格列标题
        columns = [t('file_name'), t('worksheet'), t('cell'), t('keyword'), t('content'), t('file_path')]
        for i, col in enumerate(columns):
            self.tree.heading(f"#{i+1}", text=col)
            
        # 重新创建菜单以更新语言
        self.create_menu()
    
    def update_window_title(self):
        """更新窗口标题"""
        self.root.title(t('app_title'))
    
    def search_excel(self, file_path, keywords):
        """搜索Excel文件"""
        matches = []
        keyword_counts = {keyword: 0 for keyword in keywords}
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            # 第一遍：收集所有匹配的单元格，同时统计关键词总数
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell_str = str(cell.value)
                            for keyword in keywords:
                                if keyword in cell_str:
                                    # 统计该关键词在当前单元格中的出现次数
                                    count_in_cell = cell_str.count(keyword)
                                    keyword_counts[keyword] += count_in_cell
                                    
                                    # 只存储匹配项，关键词信息稍后添加
                                    matches.append([
                                        os.path.basename(file_path),
                                        sheet_name,
                                        cell.coordinate,
                                        keyword,  # 暂时只存储关键词
                                        cell_str,
                                        file_path
                                    ])
            wb.close()
            
        except Exception as e:
            print(t('processing_error', filepath=file_path, error=str(e)))
        
        # 第二遍：为每个匹配项添加完整的关键词统计信息
        final_matches = []
        for match in matches:
            keyword = match[3]
            total_count = keyword_counts[keyword]
            updated_match = match.copy()
            updated_match[3] = f"{keyword}{t('total_occurrences', count=total_count)}"
            final_matches.append(updated_match)
        
        return final_matches, keyword_counts
    
    def search_all_excels(self, directory, keywords):
        """搜索目录中的所有Excel文件"""
        results = []
        global_keyword_counts = {keyword: 0 for keyword in keywords}  # 全局关键词统计
        
        for root, _, files in os.walk(directory):
            for file in files:
                if file.lower().endswith('.xlsx'):
                    file_path = os.path.join(root, file)
                    file_results, file_keyword_counts = self.search_excel(file_path, keywords)
                    results += file_results
                    
                    # 累加到全局统计
                    for keyword, count in file_keyword_counts.items():
                        global_keyword_counts[keyword] += count
        
        return results, global_keyword_counts
    
    def browse_directory(self):
        """浏览目录"""
        path = filedialog.askdirectory()
        if path:
            self.path_var.set(path)
            # 保存路径到配置
            self.config.set_last_search_path(path)
    
    def start_search(self):
        """开始搜索"""
        directory = self.path_var.get()
        keyword_input = self.keywords_var.get()
        
        if not os.path.exists(directory):
            messagebox.showerror(t('error'), t('invalid_path'))
            return
            
        if not keyword_input.strip():
            messagebox.showerror(t('error'), t('no_keywords'))
            return
            
        keywords = [kw.strip() for kw in keyword_input.split(",") if kw.strip()]
        results, total_stats = self.search_all_excels(directory, keywords)  # 现在直接返回统计信息
        self.current_search_results = results
        
        # 清空之前的结果
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        # 检查是否单文件搜索
        unique_files = set(row[5] for row in results)
        is_single_file = len(unique_files) == 1
        
        if is_single_file and results:
            file_path = list(unique_files)[0]
            file_name = os.path.basename(file_path)
            
            # 单文件模式：隐藏文件名和路径列
            self.tree["displaycolumns"] = (1, 2, 3, 4)
            
            for row in results:
                self.tree.insert('', 'end', values=row)
            
            stats_text = " | ".join([t('keyword_stats', keyword=kw, count=count) for kw, count in total_stats.items()])
            self.status_var.set(f"{t('search_complete', count=len(results))} | {t('file_info', filename=file_name)} | {stats_text}")
        else:
            # 多文件模式：显示所有列
            self.tree["displaycolumns"] = (0, 1, 2, 3, 4, 5)
            
            for row in results:
                self.tree.insert('', 'end', values=row)
            
            stats_text = " | ".join([t('keyword_stats', keyword=kw, count=count) for kw, count in total_stats.items()])
            self.status_var.set(f"{t('search_complete', count=len(results))} | {stats_text}")
    
    def export_csv(self):
        """导出CSV"""
        if not self.tree.get_children():
            messagebox.showinfo(t('info'), t('no_results'))
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")]
        )
        
        if file_path:
            with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                
                displayed_columns = self.tree["displaycolumns"]
                if len(displayed_columns) == 4:  # 单文件模式
                    headers = [t('worksheet'), t('cell'), t('keyword'), t('content')]
                    writer.writerow(headers)
                    
                    status_text = self.status_var.get()
                    if t('file_info', filename='') in status_text:
                        file_name = status_text.split(t('file_info', filename=''))[1].split(' |')[0]
                        writer.writerow([t('file_comment', filename=file_name), "", "", ""])
                    
                    for row_id in self.tree.get_children():
                        values = self.tree.item(row_id)["values"]
                        writer.writerow([values[1], values[2], values[3], values[4]])
                else:  # 多文件模式
                    headers = [t('file_name'), t('worksheet'), t('cell'), t('keyword'), t('content'), t('file_path')]
                    writer.writerow(headers)
                    for row_id in self.tree.get_children():
                        writer.writerow(self.tree.item(row_id)["values"])
                        
            messagebox.showinfo(t('success'), t('saved_to', filepath=file_path))
    
    def open_file(self):
        """打开选中的文件"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showinfo(t('info'), t('select_record'))
            return
        
        item = self.tree.item(selection[0])
        displayed_columns = self.tree["displaycolumns"]
        
        if len(displayed_columns) == 4:  # 单文件模式
            values = item['values']
            file_path = values[5]
        else:  # 多文件模式
            values = item['values']
            file_path = values[5]
        
        if not os.path.exists(file_path):
            messagebox.showerror(t('error'), t('file_not_exist', filepath=file_path))
            return
        
        try:
            if platform.system() == 'Windows':
                os.startfile(file_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', file_path])
            else:  # Linux
                subprocess.run(['xdg-open', file_path])
        except Exception as e:
            messagebox.showerror(t('error'), t('cannot_open_file', error=str(e)))
    
    def run(self):
        """运行应用程序"""
        self.root.mainloop()
    
    def on_closing(self):
        """窗口关闭时的处理"""
        # 保存窗口大小
        geometry = self.root.geometry()
        self.config.set_window_geometry(geometry)
        
        # 关闭窗口
        self.root.destroy()

def main():
    """主函数"""
    app = ExcelSearchApp()
    app.run()

if __name__ == "__main__":
    main()
